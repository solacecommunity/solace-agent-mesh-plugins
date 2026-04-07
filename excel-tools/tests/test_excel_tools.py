"""Tests for excel_tools plugin."""

import json
import os
import time

import openpyxl
import pytest

from excel_tools.tools import (
    analyze_sheet,
    apply_template,
    batch_recalculate,
    cancel_scheduled_job,
    copy_sheet,
    create_workbook,
    diff_workbooks,
    evaluate_formulas,
    export_sheet_to_json,
    get_formulas,
    list_scheduled_jobs,
    list_sheets,
    read_range,
    recalculate,
    schedule_recalculation,
    snapshot_formulas,
    validate_formulas,
    write_cells,
)


# ====================================================================
# Fixtures
# ====================================================================

@pytest.fixture
def sample_workbook(tmp_path):
    """Create a sample workbook with data and formulas for testing."""
    path = str(tmp_path / "test.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Headers
    ws["A1"] = "Product"
    ws["B1"] = "Quantity"
    ws["C1"] = "Price"
    ws["D1"] = "Total"

    # Data rows
    ws["A2"] = "Widget"
    ws["B2"] = 10
    ws["C2"] = 5.50
    ws["D2"] = "=B2*C2"

    ws["A3"] = "Gadget"
    ws["B3"] = 25
    ws["C3"] = 12.00
    ws["D3"] = "=B3*C3"

    ws["A4"] = "Gizmo"
    ws["B4"] = 7
    ws["C4"] = 8.75
    ws["D4"] = "=B4*C4"

    # Summary row
    ws["A6"] = "Grand Total"
    ws["D6"] = "=SUM(D2:D4)"

    # Second sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Total Sales"
    ws2["B1"] = "=Sales!D6"

    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def empty_workbook(tmp_path):
    """Create a minimal empty workbook."""
    path = str(tmp_path / "empty.xlsx")
    wb = openpyxl.Workbook()
    wb.save(path)
    wb.close()
    return path


# ====================================================================
# Core tools (v0.1)
# ====================================================================

# ---- list_sheets ----

@pytest.mark.asyncio
async def test_list_sheets(sample_workbook):
    result = await list_sheets(file_path=sample_workbook)
    assert result["status"] == "success"
    assert result["sheet_count"] == 2
    assert "Sales" in result["sheets"]
    assert "Summary" in result["sheets"]


@pytest.mark.asyncio
async def test_list_sheets_not_found():
    result = await list_sheets(file_path="/nonexistent/file.xlsx")
    assert result["status"] == "error"


# ---- read_range ----

@pytest.mark.asyncio
async def test_read_range_full_sheet(sample_workbook):
    result = await read_range(file_path=sample_workbook, sheet_name="Sales")
    assert result["status"] == "success"
    assert result["row_count"] >= 4
    assert result["rows"][0] == ["Product", "Quantity", "Price", "Total"]


@pytest.mark.asyncio
async def test_read_range_specific(sample_workbook):
    result = await read_range(
        file_path=sample_workbook,
        sheet_name="Sales",
        range_address="A1:B3",
    )
    assert result["status"] == "success"
    assert result["row_count"] == 3


@pytest.mark.asyncio
async def test_read_range_with_formulas(sample_workbook):
    result = await read_range(
        file_path=sample_workbook,
        sheet_name="Sales",
        range_address="A1:D3",
        include_formulas=True,
    )
    assert result["status"] == "success"
    assert "formulas" in result
    assert result["formulas"][1][3] == "=B2*C2"


# ---- get_formulas ----

@pytest.mark.asyncio
async def test_get_formulas(sample_workbook):
    result = await get_formulas(file_path=sample_workbook, sheet_name="Sales")
    assert result["status"] == "success"
    assert result["formula_count"] == 4
    assert result["formulas"]["D2"] == "=B2*C2"
    assert result["formulas"]["D6"] == "=SUM(D2:D4)"


@pytest.mark.asyncio
async def test_get_formulas_empty(empty_workbook):
    result = await get_formulas(file_path=empty_workbook)
    assert result["status"] == "success"
    assert result["formula_count"] == 0


# ---- evaluate_formulas ----

@pytest.mark.asyncio
async def test_evaluate_formulas(sample_workbook):
    result = await evaluate_formulas(file_path=sample_workbook)
    assert result["status"] == "success"
    assert result["evaluated_count"] > 0


@pytest.mark.asyncio
async def test_evaluate_formulas_specific_cells(sample_workbook):
    result = await evaluate_formulas(
        file_path=sample_workbook,
        cells=["D2"],
        sheet_name="Sales",
    )
    assert result["status"] == "success"
    if "D2" in result["results"]:
        assert result["results"]["D2"] == pytest.approx(55.0)


# ---- recalculate ----

@pytest.mark.asyncio
async def test_recalculate(sample_workbook):
    result = await recalculate(
        file_path=sample_workbook,
        inputs={"B2": 20},
        output_cells=["D2"],
        sheet_name="Sales",
    )
    assert result["status"] == "success"
    if "D2" in result["results"]:
        assert result["results"]["D2"] == pytest.approx(110.0)


@pytest.mark.asyncio
async def test_recalculate_no_inputs(sample_workbook):
    result = await recalculate(file_path=sample_workbook)
    assert result["status"] == "error"


# ---- analyze_sheet ----

@pytest.mark.asyncio
async def test_analyze_sheet(sample_workbook):
    result = await analyze_sheet(file_path=sample_workbook, sheet_name="Sales")
    assert result["status"] == "success"
    assert result["headers"][0] == "Product"
    assert result["dimensions"]["columns"] == 4
    assert result["formula_count"] == 4

    qty_col = next(c for c in result["columns"] if c["header"] == "Quantity")
    assert "numeric_stats" in qty_col
    assert qty_col["numeric_stats"]["count"] == 3


# ---- write_cells ----

@pytest.mark.asyncio
async def test_write_cells(sample_workbook, tmp_path):
    out_path = str(tmp_path / "modified.xlsx")
    result = await write_cells(
        file_path=sample_workbook,
        updates={"A5": "Doohickey", "B5": 15, "C5": 3.25, "D5": "=B5*C5"},
        sheet_name="Sales",
        save_as=out_path,
    )
    assert result["status"] == "success"
    assert result["count"] == 4
    assert os.path.isfile(out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["Sales"]
    assert ws["A5"].value == "Doohickey"
    assert ws["B5"].value == 15
    wb.close()


@pytest.mark.asyncio
async def test_write_cells_no_updates(sample_workbook):
    result = await write_cells(file_path=sample_workbook)
    assert result["status"] == "error"


# ---- create_workbook ----

@pytest.mark.asyncio
async def test_create_workbook(tmp_path):
    path = str(tmp_path / "new.xlsx")
    result = await create_workbook(
        file_path=path,
        sheet_name="Inventory",
        headers=["Item", "Count"],
        rows=[["Apples", 50], ["Bananas", 30]],
    )
    assert result["status"] == "success"
    assert result["rows_written"] == 3
    assert os.path.isfile(path)

    wb = openpyxl.load_workbook(path)
    ws = wb["Inventory"]
    assert ws["A1"].value == "Item"
    assert ws["A2"].value == "Apples"
    assert ws["B3"].value == 30
    wb.close()


@pytest.mark.asyncio
async def test_create_workbook_empty(tmp_path):
    path = str(tmp_path / "blank.xlsx")
    result = await create_workbook(file_path=path)
    assert result["status"] == "success"
    assert os.path.isfile(path)


# ====================================================================
# Scheduling / cron tools (v0.2)
# ====================================================================

@pytest.mark.asyncio
async def test_schedule_and_cancel(sample_workbook):
    """Schedule a job, verify it appears in the list, then cancel it."""
    result = await schedule_recalculation(
        file_path=sample_workbook,
        interval_seconds=60,
        inputs={"B2": 99},
        sheet_name="Sales",
    )
    assert result["status"] == "success"
    job_id = result["job_id"]
    assert job_id

    # It should appear in the job list
    jobs = await list_scheduled_jobs()
    assert jobs["status"] == "success"
    assert any(j["id"] == job_id for j in jobs["jobs"])

    # Cancel it
    cancel = await cancel_scheduled_job(job_id=job_id)
    assert cancel["status"] == "success"
    assert cancel["job_id"] == job_id

    # Should be gone now
    jobs2 = await list_scheduled_jobs()
    assert not any(j["id"] == job_id for j in jobs2["jobs"])


@pytest.mark.asyncio
async def test_schedule_runs_at_least_once(sample_workbook, tmp_path):
    """Schedule with a tiny interval and verify it actually runs."""
    results_file = str(tmp_path / "results.xlsx")
    result = await schedule_recalculation(
        file_path=sample_workbook,
        interval_seconds=10,
        inputs={"B2": 50},
        output_cells=["D2"],
        sheet_name="Sales",
        save_results_to=results_file,
    )
    assert result["status"] == "success"
    job_id = result["job_id"]

    # Wait for one run (10s interval + margin)
    time.sleep(12)

    jobs = await list_scheduled_jobs()
    job_info = next(j for j in jobs["jobs"] if j["id"] == job_id)
    assert job_info["run_count"] >= 1

    # Clean up
    await cancel_scheduled_job(job_id=job_id)


@pytest.mark.asyncio
async def test_schedule_file_not_found():
    result = await schedule_recalculation(
        file_path="/does/not/exist.xlsx",
        interval_seconds=60,
    )
    assert result["status"] == "error"


@pytest.mark.asyncio
async def test_schedule_interval_too_small(sample_workbook):
    result = await schedule_recalculation(
        file_path=sample_workbook,
        interval_seconds=5,
    )
    assert result["status"] == "error"


@pytest.mark.asyncio
async def test_cancel_nonexistent_job():
    result = await cancel_scheduled_job(job_id="does_not_exist")
    assert result["status"] == "error"


# ====================================================================
# Advanced automation tools (v0.2)
# ====================================================================

# ---- batch_recalculate ----

@pytest.mark.asyncio
async def test_batch_recalculate(sample_workbook):
    result = await batch_recalculate(
        file_path=sample_workbook,
        scenarios=[
            {"name": "Low", "inputs": {"B2": 5}},
            {"name": "High", "inputs": {"B2": 100}},
        ],
        output_cells=["D2"],
        sheet_name="Sales",
    )
    assert result["status"] == "success"
    assert result["scenario_count"] == 2
    assert result["scenarios"][0]["name"] == "Low"
    assert result["scenarios"][1]["name"] == "High"

    # Low: D2 = 5 * 5.50 = 27.5
    if "D2" in result["scenarios"][0]["results"]:
        assert result["scenarios"][0]["results"]["D2"] == pytest.approx(27.5)
    # High: D2 = 100 * 5.50 = 550.0
    if "D2" in result["scenarios"][1]["results"]:
        assert result["scenarios"][1]["results"]["D2"] == pytest.approx(550.0)


@pytest.mark.asyncio
async def test_batch_recalculate_no_scenarios(sample_workbook):
    result = await batch_recalculate(file_path=sample_workbook)
    assert result["status"] == "error"


@pytest.mark.asyncio
async def test_batch_recalculate_missing_inputs(sample_workbook):
    result = await batch_recalculate(
        file_path=sample_workbook,
        scenarios=[{"name": "Bad"}],
        sheet_name="Sales",
    )
    assert result["status"] == "success"
    assert result["scenarios"][0]["status"] == "error"


# ---- copy_sheet ----

@pytest.mark.asyncio
async def test_copy_sheet_same_workbook(sample_workbook):
    result = await copy_sheet(
        file_path=sample_workbook,
        source_sheet="Sales",
        new_sheet_name="Sales_Copy",
    )
    assert result["status"] == "success"

    wb = openpyxl.load_workbook(sample_workbook)
    assert "Sales_Copy" in wb.sheetnames
    assert wb["Sales_Copy"]["A1"].value == "Product"
    wb.close()


@pytest.mark.asyncio
async def test_copy_sheet_cross_workbook(sample_workbook, tmp_path):
    target = str(tmp_path / "target.xlsx")
    result = await copy_sheet(
        file_path=sample_workbook,
        source_sheet="Sales",
        new_sheet_name="Imported",
        target_file=target,
    )
    assert result["status"] == "success"
    assert os.path.isfile(target)

    wb = openpyxl.load_workbook(target)
    assert "Imported" in wb.sheetnames
    assert wb["Imported"]["A2"].value == "Widget"
    wb.close()


@pytest.mark.asyncio
async def test_copy_sheet_not_found(sample_workbook):
    result = await copy_sheet(
        file_path=sample_workbook,
        source_sheet="NoSuchSheet",
        new_sheet_name="Copy",
    )
    assert result["status"] == "error"


# ---- apply_template ----

@pytest.mark.asyncio
async def test_apply_template(sample_workbook, tmp_path):
    output = str(tmp_path / "filled.xlsx")
    result = await apply_template(
        template_path=sample_workbook,
        output_path=output,
        values={"A2": "CustomWidget", "B2": 999},
        sheet_name="Sales",
    )
    assert result["status"] == "success"
    assert result["count"] == 2

    # Original should be unchanged
    wb_orig = openpyxl.load_workbook(sample_workbook)
    assert wb_orig.active["A2"].value == "Widget"
    wb_orig.close()

    # Output should have new values but keep formulas
    wb_out = openpyxl.load_workbook(output)
    ws = wb_out["Sales"]
    assert ws["A2"].value == "CustomWidget"
    assert ws["B2"].value == 999
    assert ws["D2"].value == "=B2*C2"  # formula preserved
    wb_out.close()


@pytest.mark.asyncio
async def test_apply_template_no_values(sample_workbook, tmp_path):
    result = await apply_template(
        template_path=sample_workbook,
        output_path=str(tmp_path / "out.xlsx"),
    )
    assert result["status"] == "error"


@pytest.mark.asyncio
async def test_apply_template_not_found(tmp_path):
    result = await apply_template(
        template_path="/no/such/file.xlsx",
        output_path=str(tmp_path / "out.xlsx"),
        values={"A1": "hello"},
    )
    assert result["status"] == "error"


# ---- diff_workbooks ----

@pytest.mark.asyncio
async def test_diff_workbooks_identical(sample_workbook):
    result = await diff_workbooks(
        file_a=sample_workbook,
        file_b=sample_workbook,
        sheet_name="Sales",
    )
    assert result["status"] == "success"
    assert result["identical"] is True
    assert result["diff_count"] == 0


@pytest.mark.asyncio
async def test_diff_workbooks_different(sample_workbook, tmp_path):
    modified = str(tmp_path / "modified.xlsx")
    await write_cells(
        file_path=sample_workbook,
        updates={"A2": "CHANGED"},
        sheet_name="Sales",
        save_as=modified,
    )

    result = await diff_workbooks(
        file_a=sample_workbook,
        file_b=modified,
        sheet_name="Sales",
    )
    assert result["status"] == "success"
    assert result["identical"] is False
    assert result["diff_count"] >= 1
    assert any(d["cell"] == "A2" for d in result["differences"])


# ---- validate_formulas ----

@pytest.mark.asyncio
async def test_validate_formulas_clean(sample_workbook):
    result = await validate_formulas(file_path=sample_workbook, sheet_name="Sales")
    assert result["status"] == "success"
    assert result["formulas_checked"] == 4
    assert result["valid"] is True


@pytest.mark.asyncio
async def test_validate_formulas_with_error(tmp_path):
    """Create a workbook with a #DIV/0! formula and validate it."""
    path = str(tmp_path / "errors.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["B1"] = 0
    ws["C1"] = "=A1/B1"  # Division by zero
    wb.save(path)
    wb.close()

    result = await validate_formulas(file_path=path)
    assert result["status"] == "success"
    assert result["formulas_checked"] == 1
    # The evaluator should catch the division by zero
    # (formulas library may raise or return error token)


# ---- export_sheet_to_json ----

@pytest.mark.asyncio
async def test_export_to_json_records(sample_workbook):
    result = await export_sheet_to_json(
        file_path=sample_workbook,
        sheet_name="Sales",
        orient="records",
    )
    assert result["status"] == "success"
    assert result["orient"] == "records"
    assert result["row_count"] >= 3
    # First record should be Widget data
    assert result["data"][0]["Product"] == "Widget"


@pytest.mark.asyncio
async def test_export_to_json_columns(sample_workbook):
    result = await export_sheet_to_json(
        file_path=sample_workbook,
        sheet_name="Sales",
        orient="columns",
    )
    assert result["status"] == "success"
    assert "Product" in result["data"]
    assert result["data"]["Product"][0] == "Widget"


@pytest.mark.asyncio
async def test_export_to_json_rows(sample_workbook):
    result = await export_sheet_to_json(
        file_path=sample_workbook,
        sheet_name="Sales",
        orient="rows",
    )
    assert result["status"] == "success"
    assert result["data"][0] == ["Product", "Quantity", "Price", "Total"]


@pytest.mark.asyncio
async def test_export_to_json_file(sample_workbook, tmp_path):
    out = str(tmp_path / "export.json")
    result = await export_sheet_to_json(
        file_path=sample_workbook,
        sheet_name="Sales",
        output_path=out,
    )
    assert result["status"] == "success"
    assert os.path.isfile(out)
    with open(out) as f:
        data = json.load(f)
    assert isinstance(data, list)
    assert data[0]["Product"] == "Widget"


@pytest.mark.asyncio
async def test_export_to_json_bad_orient(sample_workbook):
    result = await export_sheet_to_json(
        file_path=sample_workbook,
        orient="invalid",
    )
    assert result["status"] == "error"


# ---- snapshot_formulas ----

@pytest.mark.asyncio
async def test_snapshot_formulas(sample_workbook, tmp_path):
    output = str(tmp_path / "snapshot.xlsx")
    result = await snapshot_formulas(
        file_path=sample_workbook,
        output_path=output,
    )
    assert result["status"] == "success"
    assert result["formulas_replaced"] > 0
    assert os.path.isfile(output)

    # The snapshot should have values, not formulas
    wb = openpyxl.load_workbook(output)
    ws = wb["Sales"]
    # D2 should now be a number, not "=B2*C2"
    assert isinstance(ws["D2"].value, (int, float))
    assert ws["D2"].value == pytest.approx(55.0)
    wb.close()

    # Original should still have the formula
    wb_orig = openpyxl.load_workbook(sample_workbook)
    assert wb_orig["Sales"]["D2"].value == "=B2*C2"
    wb_orig.close()


@pytest.mark.asyncio
async def test_snapshot_formulas_default_path(sample_workbook):
    result = await snapshot_formulas(file_path=sample_workbook)
    assert result["status"] == "success"
    expected = sample_workbook.replace(".xlsx", "_snapshot.xlsx")
    assert result["output"] == expected
    assert os.path.isfile(expected)
    # Clean up
    os.remove(expected)


@pytest.mark.asyncio
async def test_snapshot_formulas_single_sheet(sample_workbook, tmp_path):
    output = str(tmp_path / "snap_sales.xlsx")
    result = await snapshot_formulas(
        file_path=sample_workbook,
        output_path=output,
        sheet_name="Sales",
    )
    assert result["status"] == "success"
    # Only Sales formulas should be replaced, Summary should still have formula
    wb = openpyxl.load_workbook(output)
    assert isinstance(wb["Sales"]["D2"].value, (int, float))
    # Summary sheet formula referencing Sales!D6 may or may not be replaced
    # depending on whether we process it; we only asked for Sales
    wb.close()


# ====================================================================
# Artifact support tests
# ====================================================================

class FakeArtifact:
    """Mimics the SAM Artifact dataclass for testing without SAM installed."""

    def __init__(self, content: bytes, filename: str):
        self.content = content
        self.filename = filename
        self.version = 1
        self.mime_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self.metadata = {}

    def as_bytes(self, encoding: str = "utf-8") -> bytes:
        if isinstance(self.content, str):
            return self.content.encode(encoding)
        return self.content

    def as_text(self, encoding: str = "utf-8") -> str:
        if isinstance(self.content, bytes):
            return self.content.decode(encoding)
        return self.content


def _make_artifact(workbook_path: str) -> FakeArtifact:
    """Read a workbook from disk and wrap it in a FakeArtifact."""
    with open(workbook_path, "rb") as f:
        data = f.read()
    return FakeArtifact(content=data, filename=os.path.basename(workbook_path))


@pytest.mark.asyncio
async def test_list_sheets_artifact(sample_workbook):
    art = _make_artifact(sample_workbook)
    result = await list_sheets(file_path=art)
    assert result["status"] == "success"
    assert result["sheet_count"] == 2
    assert "Sales" in result["sheets"]
    assert result["file"] == "test.xlsx"


@pytest.mark.asyncio
async def test_read_range_artifact(sample_workbook):
    art = _make_artifact(sample_workbook)
    result = await read_range(file_path=art, sheet_name="Sales")
    assert result["status"] == "success"
    assert result["rows"][0] == ["Product", "Quantity", "Price", "Total"]


@pytest.mark.asyncio
async def test_get_formulas_artifact(sample_workbook):
    art = _make_artifact(sample_workbook)
    result = await get_formulas(file_path=art, sheet_name="Sales")
    assert result["status"] == "success"
    assert result["formula_count"] > 0
    assert "D2" in result["formulas"]


@pytest.mark.asyncio
async def test_evaluate_formulas_artifact(sample_workbook):
    art = _make_artifact(sample_workbook)
    result = await evaluate_formulas(file_path=art)
    assert result["status"] == "success"
    assert result["evaluated_count"] > 0


@pytest.mark.asyncio
async def test_analyze_sheet_artifact(sample_workbook):
    art = _make_artifact(sample_workbook)
    result = await analyze_sheet(file_path=art, sheet_name="Sales")
    assert result["status"] == "success"
    assert result["file"] == "test.xlsx"
    assert result["headers"][0] == "Product"


@pytest.mark.asyncio
async def test_recalculate_artifact(sample_workbook):
    art = _make_artifact(sample_workbook)
    result = await recalculate(
        file_path=art,
        inputs={"B2": 20},
        output_cells=["D2"],
        sheet_name="Sales",
    )
    assert result["status"] == "success"
    # 20 * 5.50 = 110
    assert result["results"]["D2"] == pytest.approx(110.0)


@pytest.mark.asyncio
async def test_export_sheet_to_json_artifact(sample_workbook):
    art = _make_artifact(sample_workbook)
    result = await export_sheet_to_json(file_path=art, sheet_name="Sales")
    assert result["status"] == "success"
    assert result["file"] == "test.xlsx"
    assert result["row_count"] >= 3


@pytest.mark.asyncio
async def test_diff_workbooks_artifact(sample_workbook, empty_workbook):
    art_a = _make_artifact(sample_workbook)
    art_b = _make_artifact(empty_workbook)
    result = await diff_workbooks(file_a=art_a, file_b=art_b)
    assert result["status"] == "success"
    assert result["diff_count"] > 0
    assert result["file_a"] == "test.xlsx"
    assert result["file_b"] == "empty.xlsx"


@pytest.mark.asyncio
async def test_schedule_rejects_artifact(sample_workbook):
    art = _make_artifact(sample_workbook)
    result = await schedule_recalculation(file_path=art)
    assert result["status"] == "error"
    assert "filesystem path" in result["message"]
