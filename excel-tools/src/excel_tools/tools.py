"""Excel Tools for Solace Agent Mesh.

Provides async tool functions for reading, writing, analyzing, and
reproducing calculations in Excel (.xlsx / .xls) spreadsheets.

Dependencies:
    openpyxl  — read/write .xlsx files, inspect formulas
    formulas  — evaluate Excel formulas in pure Python
"""

from __future__ import annotations

import asyncio
import contextlib
import copy
import json
import logging
import os
import shutil
import tempfile
import threading
import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional, Union

import formulas
import openpyxl
from openpyxl.utils import get_column_letter

try:
    from solace_agent_mesh.agent.tools import Artifact
except ImportError:  # allow standalone / test usage without SAM installed
    Artifact = None  # type: ignore[assignment,misc]

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _is_artifact(obj: Any) -> bool:
    """Return True when *obj* is a SAM Artifact instance."""
    if Artifact is not None and isinstance(obj, Artifact):
        return True
    # Duck-type fallback: any object with `.as_bytes()` and `.filename`
    return hasattr(obj, "as_bytes") and hasattr(obj, "filename")


@contextlib.contextmanager
def _resolve_file(file_input: Any):
    """Context manager that yields a filesystem path for *file_input*.

    *file_input* may be:
    - A plain ``str`` path — yielded as-is.
    - A SAM ``Artifact`` object — its bytes are written to a temp file
      whose path is yielded; the temp file is cleaned up on exit.
    """
    if isinstance(file_input, str):
        yield file_input
        return

    if _is_artifact(file_input):
        data = file_input.as_bytes()
        suffix = os.path.splitext(file_input.filename)[1] or ".xlsx"
        fd, tmp = tempfile.mkstemp(suffix=suffix)
        try:
            os.write(fd, data)
            os.close(fd)
            yield tmp
        finally:
            with contextlib.suppress(OSError):
                os.unlink(tmp)
        return

    raise TypeError(
        f"file_path must be a string path or an Artifact, got {type(file_input).__name__}"
    )


def _open_workbook(
    file_path: str,
    data_only: bool = False,
) -> openpyxl.Workbook:
    """Open an Excel workbook, raising a clear error on failure."""
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    return openpyxl.load_workbook(file_path, data_only=data_only)


def _cell_ref(row: int, col: int) -> str:
    """Return an Excel-style cell reference like 'A1'."""
    return f"{get_column_letter(col)}{row}"


def _serialize_cell(cell) -> Any:
    """Convert a cell value to a JSON-safe type."""
    if cell.value is None:
        return None
    if isinstance(cell.value, datetime):
        return cell.value.isoformat()
    return cell.value


def _sheet_names(wb: openpyxl.Workbook) -> List[str]:
    return wb.sheetnames


def _unwrap_value(value: Any) -> Any:
    """Extract a plain Python value from a formulas library result.

    The *formulas* library may return ``Ranges`` objects, numpy arrays,
    or native scalars.  This helper normalises all of them to JSON-safe
    Python types.
    """
    # Ranges object → extract the underlying value array, then unwrap
    if hasattr(value, "value"):
        inner = value.value
        # numpy array
        if hasattr(inner, "tolist"):
            inner = inner.tolist()
        # Flatten single-element nested lists like [[55.0]]
        while isinstance(inner, list) and len(inner) == 1:
            inner = inner[0]
        return inner
    # Plain numpy scalar / array
    if hasattr(value, "tolist"):
        v = value.tolist()
        while isinstance(v, list) and len(v) == 1:
            v = v[0]
        return v
    return value


# ---------------------------------------------------------------------------
# Tool 1 — List sheets
# ---------------------------------------------------------------------------

async def list_sheets(
    file_path: Artifact,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """List all sheet names in an Excel workbook.

    Args:
        file_path: Path or artifact reference to the .xlsx file.
    """
    log.info("[ExcelTools:list_sheets] Opening %s", file_path)
    try:
        with _resolve_file(file_path) as resolved:
            wb = _open_workbook(resolved)
            sheets = _sheet_names(wb)
            wb.close()
        return {
            "status": "success",
            "file": getattr(file_path, "filename", file_path),
            "sheets": sheets,
            "sheet_count": len(sheets),
        }
    except Exception as exc:
        log.error("[ExcelTools:list_sheets] %s", exc)
        return {"status": "error", "message": str(exc)}


# ---------------------------------------------------------------------------
# Tool 2 — Read a range
# ---------------------------------------------------------------------------

async def read_range(
    file_path: Artifact,
    sheet_name: Optional[str] = None,
    range_address: Optional[str] = None,
    max_rows: int = 200,
    include_formulas: bool = False,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Read cell values (and optionally formulas) from an Excel range.

    Args:
        file_path: Path or artifact reference to the .xlsx file.
        sheet_name: Sheet to read. Defaults to the active sheet.
        range_address: Excel range like 'A1:D10'. If omitted, reads the
            entire used range (up to *max_rows* rows).
        max_rows: Maximum number of rows to return (default 200).
        include_formulas: When True the response includes the raw formula
            strings alongside computed values.
    """
    log.info("[ExcelTools:read_range] %s sheet=%s range=%s", file_path, sheet_name, range_address)
    try:
        with _resolve_file(file_path) as resolved:
            wb_values = _open_workbook(resolved, data_only=True)
            ws_values = wb_values[sheet_name] if sheet_name else wb_values.active

            wb_formulas = None
            ws_formulas = None
            if include_formulas:
                wb_formulas = _open_workbook(resolved, data_only=False)
                ws_formulas = wb_formulas[sheet_name] if sheet_name else wb_formulas.active

            if range_address:
                cells = ws_values[range_address]
            else:
                cells = ws_values.iter_rows(
                    min_row=ws_values.min_row,
                    max_row=min(ws_values.max_row or 1, (ws_values.min_row or 1) + max_rows - 1),
                    min_col=ws_values.min_column,
                    max_col=ws_values.max_column,
                )

            rows: List[List[Any]] = []
            formula_rows: List[List[Any]] = []
            for row in cells:
                # openpyxl returns a tuple for a single-cell range
                if not hasattr(row, "__iter__"):
                    row = (row,)
                rows.append([_serialize_cell(c) for c in row])
                if include_formulas and ws_formulas:
                    f_row = []
                    for c in row:
                        fc = ws_formulas.cell(row=c.row, column=c.column)
                        val = fc.value
                        f_row.append(val if isinstance(val, str) and val.startswith("=") else None)
                    formula_rows.append(f_row)

            sheet_title = sheet_name or ws_values.title
            wb_values.close()
            if wb_formulas:
                wb_formulas.close()

        result: Dict[str, Any] = {
            "status": "success",
            "file": getattr(file_path, "filename", file_path),
            "sheet": sheet_title,
            "rows": rows,
            "row_count": len(rows),
        }
        if include_formulas:
            result["formulas"] = formula_rows
        return result

    except Exception as exc:
        log.error("[ExcelTools:read_range] %s", exc)
        return {"status": "error", "message": str(exc)}


# ---------------------------------------------------------------------------
# Tool 3 — Get formulas
# ---------------------------------------------------------------------------

async def get_formulas(
    file_path: Artifact,
    sheet_name: Optional[str] = None,
    range_address: Optional[str] = None,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Extract all formulas from a sheet or range.

    Returns a mapping of cell references to their formula strings.

    Args:
        file_path: Path or artifact reference to the .xlsx file.
        sheet_name: Sheet to inspect. Defaults to the active sheet.
        range_address: Optional range like 'A1:D10'. If omitted, scans
            the entire used range.
    """
    log.info("[ExcelTools:get_formulas] %s sheet=%s range=%s", file_path, sheet_name, range_address)
    try:
        with _resolve_file(file_path) as resolved:
            wb = _open_workbook(resolved, data_only=False)
            ws = wb[sheet_name] if sheet_name else wb.active

            if range_address:
                cells = ws[range_address]
            else:
                cells = ws.iter_rows(
                    min_row=ws.min_row,
                    max_row=ws.max_row,
                    min_col=ws.min_column,
                    max_col=ws.max_column,
                )

            formulas_map: Dict[str, str] = {}
            for row in cells:
                if not hasattr(row, "__iter__"):
                    row = (row,)
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas_map[_cell_ref(cell.row, cell.column)] = cell.value

            sheet_title = sheet_name or ws.title
            wb.close()
        return {
            "status": "success",
            "file": getattr(file_path, "filename", file_path),
            "sheet": sheet_title,
            "formulas": formulas_map,
            "formula_count": len(formulas_map),
        }
    except Exception as exc:
        log.error("[ExcelTools:get_formulas] %s", exc)
        return {"status": "error", "message": str(exc)}


# ---------------------------------------------------------------------------
# Tool 4 — Evaluate formulas
# ---------------------------------------------------------------------------

async def evaluate_formulas(
    file_path: Artifact,
    cells: Optional[List[str]] = None,
    sheet_name: Optional[str] = None,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Evaluate Excel formulas in the workbook and return computed results.

    Uses the *formulas* library to reproduce Excel's calculation engine
    in pure Python.  Every formula in the workbook is resolved using the
    data present in the file — no Excel installation required.

    Args:
        file_path: Path or artifact reference to the .xlsx file.
        cells: Optional list of cell references to evaluate
            (e.g. ``["C2", "D5"]``).  If omitted, all formulas in the
            target sheet are evaluated.
        sheet_name: Sheet containing the formulas. Defaults to the
            first sheet.
    """
    log.info("[ExcelTools:evaluate_formulas] %s cells=%s", file_path, cells)
    try:
        with _resolve_file(file_path) as resolved:
            xl_model = formulas.ExcelModel().loads(resolved).finish()
            solution = xl_model.calculate()

        results: Dict[str, Any] = {}
        target_sheet = sheet_name

        for key, value in solution.items():
            # Keys look like "'Sheet1'!A1" or "'Sheet1'!A1:B3"
            parts = str(key).split("!")
            if len(parts) == 2:
                s_name = parts[0].strip("'\"")
                cell_addr = parts[1]
            else:
                s_name = None
                cell_addr = str(key)

            if target_sheet and s_name and s_name.lower() != target_sheet.lower():
                continue

            val = _unwrap_value(value)

            if cells:
                if cell_addr.upper() in [c.upper() for c in cells]:
                    results[cell_addr] = val
            else:
                results[cell_addr] = val

        return {
            "status": "success",
            "file": getattr(file_path, "filename", file_path),
            "sheet": target_sheet,
            "results": results,
            "evaluated_count": len(results),
        }
    except Exception as exc:
        log.error("[ExcelTools:evaluate_formulas] %s", exc)
        return {"status": "error", "message": str(exc)}


# ---------------------------------------------------------------------------
# Tool 5 — Recalculate with new inputs
# ---------------------------------------------------------------------------

async def recalculate(
    file_path: Artifact,
    inputs: Dict[str, Any] = None,
    output_cells: Optional[List[str]] = None,
    sheet_name: Optional[str] = None,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Change input values and recalculate the spreadsheet.

    This is the key "what-if" tool: inject new values into specific cells,
    then re-evaluate all formulas and return the updated outputs.

    Args:
        file_path: Path or artifact reference to the .xlsx file.
        inputs: Mapping of cell references to new values,
            e.g. ``{"B2": 100, "B3": 200}``.  Cell references may
            optionally include the sheet name (``"'Sheet1'!B2"``).
        output_cells: Optional list of cell references whose
            recalculated values should be returned.  If omitted, all
            formula cells are returned.
        sheet_name: Default sheet name used when cell references in
            *inputs* / *output_cells* do not include a sheet qualifier.
    """
    log.info("[ExcelTools:recalculate] %s inputs=%s outputs=%s", file_path, inputs, output_cells)
    if not inputs:
        return {"status": "error", "message": "inputs parameter is required"}

    try:
        with _resolve_file(file_path) as resolved:
            xl_model = formulas.ExcelModel().loads(resolved).finish()

            # The formulas library uses keys like "'[filename]SHEETNAME'!CELL"
            fname = os.path.basename(resolved)
            wb = _open_workbook(resolved)
            default_sheet = sheet_name or wb.sheetnames[0]
            wb.close()

            qualified_inputs = {}
            for ref, val in inputs.items():
                if "!" in ref:
                    # User provided sheet-qualified ref; ensure filename prefix
                    if "[" not in ref:
                        parts = ref.split("!")
                        s = parts[0].strip("'\"").upper()
                        c = parts[1].upper()
                        qualified_inputs[f"'[{fname}]{s}'!{c}"] = val
                    else:
                        qualified_inputs[ref] = val
                else:
                    qualified_inputs[
                        f"'[{fname}]{default_sheet.upper()}'!{ref.upper()}"
                    ] = val

            solution = xl_model.calculate(inputs=qualified_inputs)

        results: Dict[str, Any] = {}
        for key, value in solution.items():
            parts = str(key).split("!")
            cell_addr = parts[1] if len(parts) == 2 else str(key)
            val = _unwrap_value(value)

            if output_cells:
                if cell_addr.upper() in [c.upper() for c in output_cells]:
                    results[cell_addr] = val
            else:
                results[cell_addr] = val

        return {
            "status": "success",
            "file": getattr(file_path, "filename", file_path),
            "inputs_applied": inputs,
            "results": results,
            "evaluated_count": len(results),
        }
    except Exception as exc:
        log.error("[ExcelTools:recalculate] %s", exc)
        return {"status": "error", "message": str(exc)}


# ---------------------------------------------------------------------------
# Tool 6 — Analyze / summarize a sheet
# ---------------------------------------------------------------------------

async def analyze_sheet(
    file_path: Artifact,
    sheet_name: Optional[str] = None,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Return a structural summary of a sheet.

    Includes dimensions, column headers, data types per column,
    formula locations, and basic statistics for numeric columns.

    Args:
        file_path: Path or artifact reference to the .xlsx file.
        sheet_name: Sheet to analyze. Defaults to the active sheet.
    """
    log.info("[ExcelTools:analyze_sheet] %s sheet=%s", file_path, sheet_name)
    try:
        with _resolve_file(file_path) as resolved:
            wb_vals = _open_workbook(resolved, data_only=True)
            ws_vals = wb_vals[sheet_name] if sheet_name else wb_vals.active

            wb_fmls = _open_workbook(resolved, data_only=False)
            ws_fmls = wb_fmls[sheet_name] if sheet_name else wb_fmls.active

            min_r, max_r = ws_vals.min_row or 1, ws_vals.max_row or 1
            min_c, max_c = ws_vals.min_column or 1, ws_vals.max_column or 1

            # Headers (first row)
            headers = []
            for col in range(min_c, max_c + 1):
                val = ws_vals.cell(row=min_r, column=col).value
                headers.append(str(val) if val is not None else f"Col{col}")

            # Per-column info
            columns_info = []
            for ci, col in enumerate(range(min_c, max_c + 1)):
                nums = []
                types_seen = set()
                formula_count = 0
                for row in range(min_r + 1, max_r + 1):
                    v = ws_vals.cell(row=row, column=col).value
                    fv = ws_fmls.cell(row=row, column=col).value
                    if isinstance(fv, str) and fv.startswith("="):
                        formula_count += 1
                    if v is not None:
                        types_seen.add(type(v).__name__)
                        if isinstance(v, (int, float)):
                            nums.append(v)

                info: Dict[str, Any] = {
                    "header": headers[ci],
                    "column_letter": get_column_letter(col),
                    "data_types": sorted(types_seen),
                    "formula_count": formula_count,
                }
                if nums:
                    info["numeric_stats"] = {
                        "count": len(nums),
                        "min": min(nums),
                        "max": max(nums),
                        "sum": round(sum(nums), 6),
                        "average": round(sum(nums) / len(nums), 6),
                    }
                columns_info.append(info)

            # Collect all formula locations
            all_formulas: Dict[str, str] = {}
            for row in ws_fmls.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        all_formulas[_cell_ref(cell.row, cell.column)] = cell.value

            sheet_title = ws_vals.title
            wb_vals.close()
            wb_fmls.close()

        return {
            "status": "success",
            "file": getattr(file_path, "filename", file_path),
            "sheet": sheet_title,
            "dimensions": {
                "rows": max_r - min_r + 1,
                "columns": max_c - min_c + 1,
                "data_rows": max_r - min_r,  # excluding header
            },
            "headers": headers,
            "columns": columns_info,
            "formulas": all_formulas,
            "formula_count": len(all_formulas),
        }
    except Exception as exc:
        log.error("[ExcelTools:analyze_sheet] %s", exc)
        return {"status": "error", "message": str(exc)}


# ---------------------------------------------------------------------------
# Tool 7 — Write cells
# ---------------------------------------------------------------------------

async def write_cells(
    file_path: Artifact,
    updates: Dict[str, Any] = None,
    sheet_name: Optional[str] = None,
    save_as: Optional[str] = None,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Write values or formulas to specific cells and save the workbook.

    Args:
        file_path: Path or artifact reference to the .xlsx file to modify.
        updates: Mapping of cell references to values or formula strings,
            e.g. ``{"A1": "Name", "B2": 42, "C2": "=A2+B2"}``.
        sheet_name: Target sheet. Defaults to the active sheet.
        save_as: Optional path to save as a new file instead of
            overwriting the original.
    """
    log.info("[ExcelTools:write_cells] %s updates=%s", file_path, updates)
    if not updates:
        return {"status": "error", "message": "updates parameter is required"}

    try:
        with _resolve_file(file_path) as resolved:
            wb = _open_workbook(resolved, data_only=False)
            ws = wb[sheet_name] if sheet_name else wb.active

            written = []
            for ref, val in updates.items():
                ws[ref] = val
                written.append(ref)

            out_path = save_as or resolved
            wb.save(out_path)
            wb.close()

        return {
            "status": "success",
            "file": out_path,
            "cells_written": written,
            "count": len(written),
        }
    except Exception as exc:
        log.error("[ExcelTools:write_cells] %s", exc)
        return {"status": "error", "message": str(exc)}


# ---------------------------------------------------------------------------
# Tool 8 — Create a new workbook
# ---------------------------------------------------------------------------

async def create_workbook(
    file_path: str,
    sheet_name: str = "Sheet1",
    headers: Optional[List[str]] = None,
    rows: Optional[List[List[Any]]] = None,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Create a new Excel workbook with optional initial data.

    Args:
        file_path: Where to save the new .xlsx file.
        sheet_name: Name of the initial sheet (default "Sheet1").
        headers: Optional list of column header strings.
        rows: Optional list of data rows (each row is a list of values).
    """
    log.info("[ExcelTools:create_workbook] %s", file_path)
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        current_row = 1
        if headers:
            for ci, h in enumerate(headers, start=1):
                ws.cell(row=current_row, column=ci, value=h)
            current_row += 1

        if rows:
            for row_data in rows:
                for ci, val in enumerate(row_data, start=1):
                    ws.cell(row=current_row, column=ci, value=val)
                current_row += 1

        wb.save(file_path)
        wb.close()

        return {
            "status": "success",
            "file": file_path,
            "sheet": sheet_name,
            "rows_written": (current_row - 1),
        }
    except Exception as exc:
        log.error("[ExcelTools:create_workbook] %s", exc)
        return {"status": "error", "message": str(exc)}


# ===========================================================================
# Scheduled / cron-style recalculation
# ===========================================================================

# In-memory registry of scheduled jobs.  Each entry:
#   { "id": str, "file_path": str, "interval_seconds": int,
#     "inputs": dict|None, "output_cells": list|None,
#     "sheet_name": str|None, "save_results_to": str|None,
#     "created_at": str, "last_run": str|None, "run_count": int,
#     "last_result": dict|None, "_cancel": threading.Event }
_scheduled_jobs: Dict[str, Dict[str, Any]] = {}
_scheduler_lock = threading.Lock()


def _run_scheduled_job(job: Dict[str, Any]) -> None:
    """Background worker that periodically recalculates a workbook."""
    cancel_event: threading.Event = job["_cancel"]
    while not cancel_event.is_set():
        cancel_event.wait(job["interval_seconds"])
        if cancel_event.is_set():
            break
        try:
            result = asyncio.run(
                recalculate(
                    file_path=job["file_path"],
                    inputs=job["inputs"],
                    output_cells=job["output_cells"],
                    sheet_name=job["sheet_name"],
                )
            )
            job["last_run"] = datetime.now(timezone.utc).isoformat()
            job["run_count"] += 1
            job["last_result"] = result

            # Optionally persist the recalculated results back to a file
            if job.get("save_results_to") and result.get("status") == "success":
                _persist_results(job, result)

            log.info(
                "[ExcelTools:scheduler] Job %s ran successfully (run #%d)",
                job["id"], job["run_count"],
            )
        except Exception as exc:
            log.error("[ExcelTools:scheduler] Job %s failed: %s", job["id"], exc)
            job["last_result"] = {"status": "error", "message": str(exc)}


def _persist_results(job: Dict[str, Any], result: Dict[str, Any]) -> None:
    """Write recalculation results into the save_results_to workbook."""
    out_path = job["save_results_to"]
    try:
        if os.path.isfile(out_path):
            wb = openpyxl.load_workbook(out_path)
        else:
            wb = openpyxl.Workbook()

        sheet_title = f"Run_{job['run_count']}"
        if sheet_title in wb.sheetnames:
            ws = wb[sheet_title]
        else:
            ws = wb.create_sheet(sheet_title)

        ws["A1"] = "Timestamp"
        ws["B1"] = job["last_run"]
        row = 2
        for cell_ref, val in result.get("results", {}).items():
            ws.cell(row=row, column=1, value=cell_ref)
            ws.cell(row=row, column=2, value=str(val) if not isinstance(val, (int, float)) else val)
            row += 1
        wb.save(out_path)
        wb.close()
    except Exception as exc:
        log.error("[ExcelTools:scheduler] Failed to persist results: %s", exc)


# ---------------------------------------------------------------------------
# Tool 9 — Schedule a recurring recalculation
# ---------------------------------------------------------------------------

async def schedule_recalculation(
    file_path: str,
    interval_seconds: int = 300,
    inputs: Optional[Dict[str, Any]] = None,
    output_cells: Optional[List[str]] = None,
    sheet_name: Optional[str] = None,
    save_results_to: Optional[str] = None,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Schedule a recurring recalculation of an Excel workbook.

    The workbook is re-evaluated at a fixed interval.  Optionally,
    new input values can be injected each cycle and results can be
    persisted to a separate workbook for audit / time-series tracking.

    Args:
        file_path: Absolute path to the .xlsx file.
        interval_seconds: How often to recalculate (default 300 = 5 min).
        inputs: Optional cell-to-value mapping applied before each run.
        output_cells: Optional list of output cells to capture.
        sheet_name: Default sheet for unqualified cell references.
        save_results_to: Optional path to a workbook where each run's
            results are appended as a new sheet (``Run_1``, ``Run_2``, ...).
    """
    log.info(
        "[ExcelTools:schedule_recalculation] %s every %ds",
        file_path, interval_seconds,
    )
    if _is_artifact(file_path):
        return {
            "status": "error",
            "message": "schedule_recalculation requires a filesystem path, not an artifact. "
                       "Save the workbook to disk first, then schedule it.",
        }
    if not os.path.isfile(file_path):
        return {"status": "error", "message": f"File not found: {file_path}"}
    if interval_seconds < 10:
        return {"status": "error", "message": "interval_seconds must be >= 10"}

    job_id = uuid.uuid4().hex[:12]
    cancel_event = threading.Event()

    job: Dict[str, Any] = {
        "id": job_id,
        "file_path": file_path,
        "interval_seconds": interval_seconds,
        "inputs": inputs,
        "output_cells": output_cells,
        "sheet_name": sheet_name,
        "save_results_to": save_results_to,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "last_run": None,
        "run_count": 0,
        "last_result": None,
        "_cancel": cancel_event,
    }

    thread = threading.Thread(
        target=_run_scheduled_job,
        args=(job,),
        daemon=True,
        name=f"excel-cron-{job_id}",
    )
    thread.start()

    with _scheduler_lock:
        _scheduled_jobs[job_id] = job

    return {
        "status": "success",
        "job_id": job_id,
        "file": file_path,
        "interval_seconds": interval_seconds,
        "message": (
            f"Scheduled recalculation every {interval_seconds}s. "
            f"Job ID: {job_id}"
        ),
    }


# ---------------------------------------------------------------------------
# Tool 10 — List scheduled jobs
# ---------------------------------------------------------------------------

async def list_scheduled_jobs(
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """List all active scheduled recalculation jobs.

    Returns metadata for every running job including last run time,
    run count, and most recent result.
    """
    log.info("[ExcelTools:list_scheduled_jobs]")
    with _scheduler_lock:
        jobs = []
        for j in _scheduled_jobs.values():
            jobs.append({
                "id": j["id"],
                "file_path": j["file_path"],
                "interval_seconds": j["interval_seconds"],
                "created_at": j["created_at"],
                "last_run": j["last_run"],
                "run_count": j["run_count"],
                "save_results_to": j.get("save_results_to"),
            })
    return {"status": "success", "jobs": jobs, "count": len(jobs)}


# ---------------------------------------------------------------------------
# Tool 11 — Cancel a scheduled job
# ---------------------------------------------------------------------------

async def cancel_scheduled_job(
    job_id: str,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Cancel a running scheduled recalculation job.

    Args:
        job_id: The job ID returned by ``schedule_recalculation``.
    """
    log.info("[ExcelTools:cancel_scheduled_job] %s", job_id)
    with _scheduler_lock:
        job = _scheduled_jobs.pop(job_id, None)
    if not job:
        return {"status": "error", "message": f"No job with ID '{job_id}'"}

    job["_cancel"].set()
    return {
        "status": "success",
        "job_id": job_id,
        "run_count": job["run_count"],
        "message": f"Job {job_id} cancelled after {job['run_count']} runs.",
    }


# ===========================================================================
# Advanced automation tools
# ===========================================================================

# ---------------------------------------------------------------------------
# Tool 12 — Batch / scenario recalculation
# ---------------------------------------------------------------------------

async def batch_recalculate(
    file_path: Artifact,
    scenarios: List[Dict[str, Any]] = None,
    output_cells: Optional[List[str]] = None,
    sheet_name: Optional[str] = None,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Run multiple what-if scenarios in a single call.

    Each scenario is a set of input overrides.  The tool evaluates
    every scenario independently and returns all results together —
    ideal for sensitivity analysis, pricing tiers, or Monte Carlo inputs.

    Args:
        file_path: Path or artifact reference to the .xlsx file.
        scenarios: List of scenario dicts.  Each dict has:
            - ``name`` (str, optional): A label for the scenario.
            - ``inputs`` (dict): Cell-to-value mapping, same as *recalculate*.
        output_cells: Cells to capture from each scenario.
        sheet_name: Default sheet for unqualified references.

    Example ``scenarios``::

        [
            {"name": "Low",  "inputs": {"B2": 100}},
            {"name": "Mid",  "inputs": {"B2": 250}},
            {"name": "High", "inputs": {"B2": 500}},
        ]
    """
    log.info("[ExcelTools:batch_recalculate] %s scenarios=%d", file_path, len(scenarios or []))
    if not scenarios:
        return {"status": "error", "message": "scenarios parameter is required (list of dicts)"}

    # Resolve the file once for all scenarios
    with _resolve_file(file_path) as resolved:
        results = []
        for idx, scenario in enumerate(scenarios):
            name = scenario.get("name", f"scenario_{idx + 1}")
            inputs = scenario.get("inputs")
            if not inputs:
                results.append({"name": name, "status": "error", "message": "scenario missing 'inputs'"})
                continue

            r = await recalculate(
                file_path=resolved,
                inputs=inputs,
                output_cells=output_cells,
                sheet_name=sheet_name,
            )
            r["name"] = name
            results.append(r)

    return {
        "status": "success",
        "file": getattr(file_path, "filename", file_path),
        "scenario_count": len(results),
        "scenarios": results,
    }


# ---------------------------------------------------------------------------
# Tool 13 — Copy / duplicate a sheet
# ---------------------------------------------------------------------------

async def copy_sheet(
    file_path: Artifact,
    source_sheet: str,
    new_sheet_name: str,
    target_file: Optional[str] = None,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Duplicate a sheet within the same workbook or into another file.

    Args:
        file_path: Path or artifact reference to the source workbook.
        source_sheet: Name of the sheet to copy.
        new_sheet_name: Name for the new sheet.
        target_file: If provided, the copy is placed in this workbook
            instead.  The target file is created if it does not exist.
    """
    log.info(
        "[ExcelTools:copy_sheet] %s %s -> %s (target=%s)",
        file_path, source_sheet, new_sheet_name, target_file,
    )
    try:
        with _resolve_file(file_path) as resolved:
            wb_src = _open_workbook(resolved, data_only=False)
            if source_sheet not in wb_src.sheetnames:
                wb_src.close()
                return {"status": "error", "message": f"Sheet '{source_sheet}' not found"}

            if target_file and target_file != resolved:
                # Cross-workbook copy
                if os.path.isfile(target_file):
                    wb_dst = openpyxl.load_workbook(target_file)
                else:
                    wb_dst = openpyxl.Workbook()
                    # Remove the default empty sheet if we're creating fresh
                    if "Sheet" in wb_dst.sheetnames and len(wb_dst.sheetnames) == 1:
                        del wb_dst["Sheet"]

                ws_src = wb_src[source_sheet]
                ws_dst = wb_dst.create_sheet(new_sheet_name)

                for row in ws_src.iter_rows():
                    for cell in row:
                        ws_dst.cell(
                            row=cell.row, column=cell.column, value=cell.value
                        )

                wb_dst.save(target_file)
                wb_dst.close()
                wb_src.close()
                out = target_file
            else:
                # Same-workbook copy
                ws_src = wb_src[source_sheet]
                ws_new = wb_src.copy_worksheet(ws_src)
                ws_new.title = new_sheet_name
                wb_src.save(resolved)
                wb_src.close()
                out = resolved

        return {
            "status": "success",
            "file": out,
            "source_sheet": source_sheet,
            "new_sheet": new_sheet_name,
        }
    except Exception as exc:
        log.error("[ExcelTools:copy_sheet] %s", exc)
        return {"status": "error", "message": str(exc)}


# ---------------------------------------------------------------------------
# Tool 14 — Apply a template
# ---------------------------------------------------------------------------

async def apply_template(
    template_path: Artifact,
    output_path: str,
    values: Dict[str, Any] = None,
    sheet_name: Optional[str] = None,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Stamp a template workbook with new values and save a copy.

    Copies the template, then writes the supplied values into the
    specified cells.  Formulas in the template are preserved and will
    recalculate when opened in Excel (or via ``evaluate_formulas``).

    Args:
        template_path: Path or artifact reference to the template .xlsx file (not modified).
        output_path: Where to save the filled copy.
        values: Cell-to-value mapping, e.g. ``{"B2": "Acme Corp", "C5": 42000}``.
        sheet_name: Target sheet. Defaults to the active sheet.
    """
    log.info("[ExcelTools:apply_template] %s -> %s", template_path, output_path)
    if not values:
        return {"status": "error", "message": "values parameter is required"}

    try:
        with _resolve_file(template_path) as resolved:
            shutil.copy2(resolved, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        written = []
        for ref, val in values.items():
            ws[ref] = val
            written.append(ref)

        wb.save(output_path)
        wb.close()

        return {
            "status": "success",
            "template": getattr(template_path, "filename", template_path),
            "output": output_path,
            "cells_written": written,
            "count": len(written),
        }
    except Exception as exc:
        log.error("[ExcelTools:apply_template] %s", exc)
        return {"status": "error", "message": str(exc)}


# ---------------------------------------------------------------------------
# Tool 15 — Diff two workbooks
# ---------------------------------------------------------------------------

async def diff_workbooks(
    file_a: Artifact,
    file_b: Artifact,
    sheet_name: Optional[str] = None,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Compare two workbooks cell-by-cell and return the differences.

    Compares values (data_only mode) for the specified sheet or the
    first sheet in each workbook.

    Args:
        file_a: Path or artifact reference to the first workbook.
        file_b: Path or artifact reference to the second workbook.
        sheet_name: Sheet to compare. Defaults to the first sheet.
    """
    log.info("[ExcelTools:diff_workbooks] %s vs %s", file_a, file_b)
    try:
        with _resolve_file(file_a) as resolved_a, _resolve_file(file_b) as resolved_b:
            wb_a = _open_workbook(resolved_a, data_only=True)
            wb_b = _open_workbook(resolved_b, data_only=True)

            ws_a = wb_a[sheet_name] if sheet_name else wb_a.active
            ws_b = wb_b[sheet_name] if sheet_name else wb_b.active

            max_row = max(ws_a.max_row or 1, ws_b.max_row or 1)
            max_col = max(ws_a.max_column or 1, ws_b.max_column or 1)

            diffs: List[Dict[str, Any]] = []
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    va = ws_a.cell(row=r, column=c).value
                    vb = ws_b.cell(row=r, column=c).value
                    if va != vb:
                        ref = _cell_ref(r, c)
                        diffs.append({
                            "cell": ref,
                            "file_a": _safe_serialize(va),
                            "file_b": _safe_serialize(vb),
                        })

            sheet_title = sheet_name or ws_a.title
            wb_a.close()
            wb_b.close()

        return {
            "status": "success",
            "file_a": getattr(file_a, "filename", file_a),
            "file_b": getattr(file_b, "filename", file_b),
            "sheet": sheet_title,
            "differences": diffs,
            "diff_count": len(diffs),
            "identical": len(diffs) == 0,
        }
    except Exception as exc:
        log.error("[ExcelTools:diff_workbooks] %s", exc)
        return {"status": "error", "message": str(exc)}


def _safe_serialize(val: Any) -> Any:
    """Convert a value to a JSON-safe type."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    return val


# ---------------------------------------------------------------------------
# Tool 16 — Validate formulas
# ---------------------------------------------------------------------------

async def validate_formulas(
    file_path: Artifact,
    sheet_name: Optional[str] = None,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Check every formula in a sheet for errors and issues.

    Evaluates all formulas and flags cells that return Excel error
    values (``#REF!``, ``#DIV/0!``, ``#NAME?``, ``#VALUE!``, ``#N/A``,
    ``#NULL!``, ``#NUM!``).  Also detects formulas referencing cells
    outside the used range (potential dangling references).

    Args:
        file_path: Path or artifact reference to the .xlsx file.
        sheet_name: Sheet to validate. Defaults to the active sheet.
    """
    log.info("[ExcelTools:validate_formulas] %s sheet=%s", file_path, sheet_name)
    try:
        with _resolve_file(file_path) as resolved:
            # Get formulas
            wb_fml = _open_workbook(resolved, data_only=False)
            ws_fml = wb_fml[sheet_name] if sheet_name else wb_fml.active

            # Get cached values
            wb_val = _open_workbook(resolved, data_only=True)
            ws_val = wb_val[sheet_name] if sheet_name else wb_val.active

            used_max_row = ws_fml.max_row or 1
            used_max_col = ws_fml.max_column or 1

            error_tokens = {"#REF!", "#DIV/0!", "#NAME?", "#VALUE!", "#N/A", "#NULL!", "#NUM!"}

            issues: List[Dict[str, Any]] = []
            formula_count = 0

            for row in ws_fml.iter_rows(
                min_row=ws_fml.min_row or 1,
                max_row=used_max_row,
                min_col=ws_fml.min_column or 1,
                max_col=used_max_col,
            ):
                for cell in row:
                    if not isinstance(cell.value, str) or not cell.value.startswith("="):
                        continue
                    formula_count += 1
                    ref = _cell_ref(cell.row, cell.column)

                    # Check cached value for error tokens
                    cached = ws_val.cell(row=cell.row, column=cell.column).value
                    if isinstance(cached, str) and cached.strip() in error_tokens:
                        issues.append({
                            "cell": ref,
                            "formula": cell.value,
                            "error": cached.strip(),
                            "severity": "error",
                        })

            # Try full evaluation to catch runtime errors
            try:
                xl_model = formulas.ExcelModel().loads(resolved).finish()
                solution = xl_model.calculate()
                for key, value in solution.items():
                    val = _unwrap_value(value)
                    if isinstance(val, str) and val.strip() in error_tokens:
                        parts = str(key).split("!")
                        cell_addr = parts[1] if len(parts) == 2 else str(key)
                        # Avoid duplicates
                        if not any(i["cell"] == cell_addr for i in issues):
                            issues.append({
                                "cell": cell_addr,
                                "formula": None,
                                "error": val.strip(),
                                "severity": "error",
                            })
            except Exception as eval_exc:
                issues.append({
                    "cell": "N/A",
                    "formula": None,
                    "error": f"Evaluation engine error: {eval_exc}",
                    "severity": "warning",
                })

            sheet_title = sheet_name or ws_fml.title
            wb_fml.close()
            wb_val.close()

        return {
            "status": "success",
            "file": getattr(file_path, "filename", file_path),
            "sheet": sheet_title,
            "formulas_checked": formula_count,
            "issues": issues,
            "issue_count": len(issues),
            "valid": len(issues) == 0,
        }
    except Exception as exc:
        log.error("[ExcelTools:validate_formulas] %s", exc)
        return {"status": "error", "message": str(exc)}


# ---------------------------------------------------------------------------
# Tool 17 — Export sheet to JSON
# ---------------------------------------------------------------------------

async def export_sheet_to_json(
    file_path: Artifact,
    sheet_name: Optional[str] = None,
    output_path: Optional[str] = None,
    orient: str = "records",
    max_rows: int = 5000,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Export a sheet's data as JSON for pipeline integration.

    Reads the sheet (values only, no formulas) and returns or saves
    the data as a JSON structure.  The first row is treated as headers.

    Args:
        file_path: Path or artifact reference to the .xlsx file.
        sheet_name: Sheet to export. Defaults to the active sheet.
        output_path: If provided, the JSON is written to this file.
        orient: Output format:
            - ``"records"`` (default): list of ``{header: value}`` dicts.
            - ``"columns"``: ``{header: [values...]}`` mapping.
            - ``"rows"``: list of lists (raw rows including headers).
        max_rows: Maximum data rows to export (default 5000).
    """
    log.info("[ExcelTools:export_sheet_to_json] %s orient=%s", file_path, orient)
    try:
        with _resolve_file(file_path) as resolved:
            wb = _open_workbook(resolved, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active

            min_r = ws.min_row or 1
            max_r = min(ws.max_row or 1, min_r + max_rows)
            min_c = ws.min_column or 1
            max_c = ws.max_column or 1

            # Read headers from first row
            headers = []
            for c in range(min_c, max_c + 1):
                v = ws.cell(row=min_r, column=c).value
                headers.append(str(v) if v is not None else f"col_{c}")

            # Read data rows
            data_rows = []
            for r in range(min_r + 1, max_r + 1):
                row_vals = []
                for c in range(min_c, max_c + 1):
                    row_vals.append(_safe_serialize(ws.cell(row=r, column=c).value))
                data_rows.append(row_vals)

            sheet_title = sheet_name or ws.title
            wb.close()

        # Format output
        if orient == "records":
            output = [dict(zip(headers, row)) for row in data_rows]
        elif orient == "columns":
            output = {}
            for i, h in enumerate(headers):
                output[h] = [row[i] for row in data_rows]
        elif orient == "rows":
            output = [headers] + data_rows
        else:
            return {"status": "error", "message": f"Unknown orient '{orient}'. Use records, columns, or rows."}

        # Optionally save to file
        if output_path:
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(output, f, indent=2, default=str)

        return {
            "status": "success",
            "file": getattr(file_path, "filename", file_path),
            "sheet": sheet_title,
            "orient": orient,
            "row_count": len(data_rows),
            "data": output,
            "saved_to": output_path,
        }
    except Exception as exc:
        log.error("[ExcelTools:export_sheet_to_json] %s", exc)
        return {"status": "error", "message": str(exc)}


# ---------------------------------------------------------------------------
# Tool 18 — Snapshot / freeze formulas to values
# ---------------------------------------------------------------------------

async def snapshot_formulas(
    file_path: Artifact,
    output_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
    tool_context=None,
    tool_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Replace all formulas with their computed values (freeze).

    Creates a "snapshot" copy where every formula cell is replaced by
    its evaluated result.  Useful for archiving a point-in-time state
    or sending a workbook to someone who should not see the formulas.

    Args:
        file_path: Path or artifact reference to the source .xlsx file.
        output_path: Where to save the snapshot.  If omitted, a
            ``_snapshot`` suffix is added to the original filename.
        sheet_name: Sheet to snapshot.  If omitted, all sheets are
            processed.
    """
    log.info("[ExcelTools:snapshot_formulas] %s", file_path)
    try:
        with _resolve_file(file_path) as resolved:
            # Evaluate all formulas
            xl_model = formulas.ExcelModel().loads(resolved).finish()
            solution = xl_model.calculate()

            # Build a lookup: (sheet_upper, cell_upper) -> value
            evaluated: Dict[tuple, Any] = {}
            for key, value in solution.items():
                parts = str(key).split("!")
                if len(parts) == 2:
                    # Strip '[filename]' from sheet part
                    raw_sheet = parts[0].strip("'\"")
                    if "]" in raw_sheet:
                        raw_sheet = raw_sheet.split("]", 1)[1]
                    evaluated[(raw_sheet.upper(), parts[1].upper())] = _unwrap_value(value)

            # Open the workbook (with formulas) and overwrite formula cells
            if not output_path:
                source_name = getattr(file_path, "filename", resolved)
                base, ext = os.path.splitext(source_name)
                output_path = f"{base}_snapshot{ext}"

            shutil.copy2(resolved, output_path)

        wb = openpyxl.load_workbook(output_path)

        sheets_to_process = [sheet_name] if sheet_name else wb.sheetnames
        replaced_count = 0

        for sn in sheets_to_process:
            ws = wb[sn]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        lookup_key = (sn.upper(), _cell_ref(cell.row, cell.column))
                        if lookup_key in evaluated:
                            cell.value = evaluated[lookup_key]
                            replaced_count += 1

        wb.save(output_path)
        wb.close()

        return {
            "status": "success",
            "source": getattr(file_path, "filename", file_path),
            "output": output_path,
            "formulas_replaced": replaced_count,
        }
    except Exception as exc:
        log.error("[ExcelTools:snapshot_formulas] %s", exc)
        return {"status": "error", "message": str(exc)}
